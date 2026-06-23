from flask import Flask, request, redirect, render_template, send_file, session
from flask_socketio import SocketIO
import sqlite3
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from io import BytesIO

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")

app.secret_key = os.environ.get("SECRET_KEY", "clave-local")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "1234")
COLAB_PASSWORD = os.environ.get("COLAB_PASSWORD", "2222")

DB_PATH = os.environ.get("DB_PATH", "scores.db")
UPLOAD_FOLDER = "uploads"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def categoria_por_hcp(hcp):
    if 0 <= hcp <= 12:
        return "0 a 12"
    elif 13 <= hcp <= 22:
        return "13 a 22"
    elif 23 <= hcp <= 36:
        return "23 a 36"
    return "Sin categoría"

def crear_o_actualizar_jugador_ranking(nombre, handicap, categoria_anual):
    con = db()

    jugador = con.execute("""
        SELECT *
        FROM jugadores
        WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(?))
    """, (nombre,)).fetchone()

    if jugador:
        jugador_id = jugador["id"]

        con.execute("""
            UPDATE jugadores
            SET handicap = ?, categoria_anual = ?
            WHERE id = ?
        """, (handicap, categoria_anual, jugador_id))

    else:
        con.execute("""
            INSERT INTO jugadores (nombre, handicap, categoria_anual)
            VALUES (?, ?, ?)
        """, (nombre, handicap, categoria_anual))

        jugador_id = con.execute("""
            SELECT last_insert_rowid()
        """).fetchone()[0]

        matricula = f"MGT-{jugador_id:04d}"

        con.execute("""
            UPDATE jugadores
            SET matricula = ?
            WHERE id = ?
        """, (matricula, jugador_id))

    con.commit()
    con.close()

    return jugador_id

def guardar_punto_ranking(jugador_id, tipo_ranking, numero_fecha, puntos):
    if puntos in [None, "", "-"]:
        return

    con = db()

    con.execute("""
        INSERT INTO ranking_fechas
        (jugador_id, tipo_ranking, numero_fecha, puntos)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(jugador_id, tipo_ranking, numero_fecha)
        DO UPDATE SET puntos = excluded.puntos
    """, (
        jugador_id,
        tipo_ranking,
        numero_fecha,
        float(puntos)
    ))

    con.commit()
    con.close()

def db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    con = db()

    con.execute("""
        CREATE TABLE IF NOT EXISTS jugadores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE,
            handicap INTEGER NOT NULL
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS tarjetas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jugador_id INTEGER NOT NULL UNIQUE,
            nombre TEXT NOT NULL,
            handicap INTEGER NOT NULL,
            ida INTEGER NOT NULL,
            vuelta INTEGER NOT NULL,
            gross INTEGER NOT NULL,
            neto INTEGER NOT NULL,
            categoria TEXT NOT NULL
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS configuracion (
            id INTEGER PRIMARY KEY,
            titulo TEXT NOT NULL,
            subtitulo TEXT NOT NULL,
            subtitulo2 TEXT NOT NULL,
            logo TEXT NOT NULL,
            fondo TEXT NOT NULL
        )
    """)

    existe = con.execute("""
        SELECT id
        FROM configuracion
        WHERE id = 1
    """).fetchone()

    if not existe:
        con.execute("""
            INSERT INTO configuracion
            (id, titulo, subtitulo, subtitulo2, logo, fondo)
            VALUES
            (1, 'Torneo de Golf', 'Tabla de posiciones por categoría', 'Resultados oficiales', '', '')
        """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS jugadores_equipos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            equipo TEXT NOT NULL,
            comodin INTEGER NOT NULL DEFAULT 0
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS matches_equipos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_match INTEGER NOT NULL,
            jugador_team22_id INTEGER NOT NULL,
            jugador_aguilas_id INTEGER NOT NULL,
            puntos_partido_team22 REAL DEFAULT 0,
            puntos_partido_aguilas REAL DEFAULT 0,
            puntos_tabla_team22 REAL DEFAULT 0,
            puntos_tabla_aguilas REAL DEFAULT 0,
            resultado TEXT DEFAULT 'Pendiente',
            resultado_cargado INTEGER NOT NULL DEFAULT 0
        )
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS puntos_equipos (
            equipo TEXT PRIMARY KEY,
            puntos_previos REAL NOT NULL DEFAULT 0
        )
    """)

    for equipo in ["Team 22", "Águilas"]:
        existe = con.execute("""
            SELECT equipo
            FROM puntos_equipos
            WHERE equipo = ?
        """, (equipo,)).fetchone()

        if not existe:
            con.execute("""
                INSERT INTO puntos_equipos (equipo, puntos_previos)
                VALUES (?, 0)
            """, (equipo,))

    con.execute("""
        CREATE TABLE IF NOT EXISTS premios_especiales (
            id INTEGER PRIMARY KEY,
            tipo TEXT NOT NULL,
            categoria TEXT NOT NULL,
            jugador_id INTEGER
        )
    """)

    premios_base = [
        (1, "Approach", "0 a 18"),
        (2, "Approach", "19 a 36"),
        (3, "Long Drive", "0 a 18"),
        (4, "Long Drive", "19 a 36")
    ]

    for premio in premios_base:
        existe = con.execute("""
            SELECT id
            FROM premios_especiales
            WHERE id = ?
        """, (premio[0],)).fetchone()

        if not existe:
            con.execute("""
                INSERT INTO premios_especiales
                (id, tipo, categoria, jugador_id)
                VALUES (?, ?, ?, NULL)
            """, premio)

    con.execute("""
    CREATE TABLE IF NOT EXISTS ranking_anual (
        jugador_id INTEGER PRIMARY KEY,
        nombre TEXT NOT NULL,
        handicap INTEGER NOT NULL,
        categoria TEXT NOT NULL,

        puntos_categoria_previos REAL NOT NULL DEFAULT 0,
        puntos_categoria_actual REAL NOT NULL DEFAULT 0,
        puntos_categoria_total REAL NOT NULL DEFAULT 0,

        puntos_general_previos REAL NOT NULL DEFAULT 0,
        puntos_general_actual REAL NOT NULL DEFAULT 0,
        puntos_general_total REAL NOT NULL DEFAULT 0
        )
    """)        
    con.execute("""
        CREATE TABLE IF NOT EXISTS ranking_fechas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jugador_id INTEGER NOT NULL,
            tipo_ranking TEXT NOT NULL,
            numero_fecha INTEGER NOT NULL,
            puntos REAL NOT NULL DEFAULT 0,
            UNIQUE(jugador_id, tipo_ranking, numero_fecha)
        )
    """)
    con.commit()
    con.close()

def asignar_matriculas():
    con = db()

    jugadores = con.execute("""
        SELECT id, matricula
        FROM jugadores
        ORDER BY id
    """).fetchall()

    for jugador in jugadores:

        if jugador["matricula"]:
            continue

        matricula = f"MGT-{jugador['id']:04d}"

        con.execute("""
            UPDATE jugadores
            SET matricula = ?
            WHERE id = ?
        """, (
            matricula,
            jugador["id"]
        ))

    con.commit()
    con.close()

def admin_logueado():
    return session.get("admin") is True

def colaborador_logueado():
    return session.get("colaborador") is True

def admin_o_colaborador():
    return session.get("admin") is True or session.get("colaborador") is True

def redireccion_post_carga(ok):
    if session.get("admin") is True:
        return redirect(f"/admin?ok={ok}")

    if session.get("colaborador") is True:
        return redirect(f"/colaborador?ok={ok}")

    return redirect("/login")

def obtener_configuracion():
    con = db()

    config = con.execute("""
        SELECT *
        FROM configuracion
        WHERE id = 1
    """).fetchone()

    con.close()

    if config:
        return dict(config)

    return {
        "titulo": "Torneo de Golf",
        "subtitulo": "",
        "subtitulo2": "",
        "logo": "",
        "fondo": ""
    }

def puntos_por_posicion(posicion):
    puntos = {
        1: 20,
        2: 16,
        3: 13,
        4: 10,
        5: 8,
        6: 6,
        7: 4,
        8: 3,
        9: 2,
        10: 1
    }

    return puntos.get(posicion, 0)

def agregar_puntos(lista):
    lista_ordenada = sorted(
        lista,
        key=lambda j: (
            float(j["neto"]),
            float(j["vuelta"]) - (float(j["handicap"]) / 2)
        )
    )

    resultado = []
    neto_anterior = None
    ranking_puntos = 0

    for jugador in lista_ordenada:
        jugador = dict(jugador)

        if float(jugador["neto"]) != neto_anterior:
            ranking_puntos += 1
            neto_anterior = float(jugador["neto"])

        jugador["puntos"] = puntos_por_posicion(ranking_puntos)

        resultado.append(jugador)

    return resultado

def obtener_puntos_previos():
    con = db()

    datos = con.execute("""
        SELECT *
        FROM puntos_equipos
    """).fetchall()

    con.close()

    puntos = {
        "Team 22": 0,
        "Águilas": 0
    }

    for d in datos:
        puntos[d["equipo"]] = float(d["puntos_previos"])

    return puntos

def obtener_categorias():
    con = db()
    categorias = {}

    for cat in ["0 a 12", "13 a 22", "23 a 36"]:
        lista = con.execute("""
            SELECT *
            FROM tarjetas
            WHERE categoria = ?
            ORDER BY neto ASC
        """, (cat,)).fetchall()

        categorias[cat] = agregar_puntos(lista)

    con.close()
    return categorias

def obtener_general():
    con = db()

    lista = con.execute("""
        SELECT *
        FROM tarjetas
        ORDER BY neto ASC
    """).fetchall()

    con.close()

    return agregar_puntos(lista)

def inicializar_ranking_anual():
    con = db()

    jugadores = con.execute("""
        SELECT *
        FROM jugadores
    """).fetchall()

    for jugador in jugadores:

        existe = con.execute("""
            SELECT jugador_id
            FROM ranking_anual
            WHERE jugador_id = ?
        """, (jugador["id"],)).fetchone()

        if not existe:

            categoria = categoria_por_hcp(
                int(jugador["handicap"])
            )

            con.execute("""
                INSERT INTO ranking_anual
                (
                    jugador_id,
                    nombre,
                    handicap,
                    categoria
                )
                VALUES (?, ?, ?, ?)
            """, (
                jugador["id"],
                jugador["nombre"],
                jugador["handicap"],
                categoria
            ))

    con.commit()
    con.close()    

def actualizar_puntos_actuales_ranking():
    con = db()

    con.execute("""
        UPDATE ranking_anual
        SET
            puntos_categoria_actual = 0,
            puntos_categoria_total = puntos_categoria_previos,
            puntos_general_actual = 0,
            puntos_general_total = puntos_general_previos
    """)

    categorias = obtener_categorias()

    for categoria, jugadores in categorias.items():
        for jugador in jugadores:
            con.execute("""
                UPDATE ranking_anual
                SET
                    puntos_categoria_actual = ?,
                    puntos_categoria_total = puntos_categoria_previos + ?
                WHERE jugador_id = ?
            """, (
                jugador["puntos"],
                jugador["puntos"],
                jugador["jugador_id"]
            ))

    general = obtener_general()

    for jugador in general:
        con.execute("""
            UPDATE ranking_anual
            SET
                puntos_general_actual = ?,
                puntos_general_total = puntos_general_previos + ?
            WHERE jugador_id = ?
        """, (
            jugador["puntos"],
            jugador["puntos"],
            jugador["jugador_id"]
        ))

    con.commit()
    con.close()

def obtener_ranking_anual():
    actualizar_puntos_actuales_ranking()

    con = db()

    rankings = {}

    for categoria in ["0 a 12", "13 a 22", "23 a 36"]:
        rankings[categoria] = con.execute("""
            SELECT *
            FROM ranking_anual
            WHERE categoria = ?
            ORDER BY puntos_categoria_total DESC, nombre ASC
        """, (categoria,)).fetchall()

    general = con.execute("""
        SELECT *
        FROM ranking_anual
        ORDER BY puntos_general_total DESC, nombre ASC
    """).fetchall()

    con.close()

    return {
        "categorias": rankings,
        "general": general
    }

def obtener_ultima_fecha_ranking():
    con = db()

    ultima = con.execute("""
        SELECT MAX(numero_fecha) AS ultima
        FROM ranking_fechas
    """).fetchone()

    con.close()

    if ultima and ultima["ultima"]:
        return int(ultima["ultima"])

    return 0

def obtener_proxima_fecha_ranking():
    return obtener_ultima_fecha_ranking() + 1

def obtener_ranking_por_fechas():
    con = db()

    fechas = con.execute("""
        SELECT DISTINCT numero_fecha
        FROM ranking_fechas
        ORDER BY numero_fecha ASC
    """).fetchall()

    fechas = [f["numero_fecha"] for f in fechas]

    jugadores = con.execute("""
        SELECT id, nombre, handicap, categoria_anual
        FROM jugadores
        ORDER BY nombre ASC
    """).fetchall()

    datos = con.execute("""
        SELECT jugador_id, tipo_ranking, numero_fecha, puntos
        FROM ranking_fechas
    """).fetchall()

    con.close()

    mapa = {}

    for d in datos:
        clave = (
            d["jugador_id"],
            d["tipo_ranking"],
            d["numero_fecha"]
        )
        mapa[clave] = d["puntos"]

    rankings = {
        "categorias": {
            "0 a 12": [],
            "13 a 22": [],
            "23 a 36": []
        },
        "general": [],
        "fechas": fechas
    }

    for jugador in jugadores:
        categoria = jugador["categoria_anual"]

        fila_categoria = {
            "nombre": jugador["nombre"],
            "handicap": jugador["handicap"],
            "fechas": {},
            "total": 0
        }

        fila_general = {
            "nombre": jugador["nombre"],
            "handicap": jugador["handicap"],
            "fechas": {},
            "total": 0
        }

        for fecha in fechas:
            puntos_cat = mapa.get((jugador["id"], "categoria", fecha))
            puntos_gen = mapa.get((jugador["id"], "general", fecha))

            fila_categoria["fechas"][fecha] = puntos_cat if puntos_cat is not None else "-"
            fila_general["fechas"][fecha] = puntos_gen if puntos_gen is not None else "-"

            if puntos_cat is not None:
                fila_categoria["total"] += puntos_cat

            if puntos_gen is not None:
                fila_general["total"] += puntos_gen

        if categoria in rankings["categorias"]:
            rankings["categorias"][categoria].append(fila_categoria)

        rankings["general"].append(fila_general)

    for categoria in rankings["categorias"]:
        rankings["categorias"][categoria].sort(
            key=lambda x: (-x["total"], x["nombre"])
        )

    rankings["general"].sort(
        key=lambda x: (-x["total"], x["nombre"])
    )

    return rankings

def obtener_jugadores_premios():
    con = db()

    jugadores_0_18 = con.execute("""
        SELECT *
        FROM jugadores
        WHERE handicap BETWEEN 0 AND 18
        ORDER BY nombre ASC
    """).fetchall()

    jugadores_19_36 = con.execute("""
        SELECT *
        FROM jugadores
        WHERE handicap BETWEEN 19 AND 36
        ORDER BY nombre ASC
    """).fetchall()

    con.close()

    return {
        "0 a 18": jugadores_0_18,
        "19 a 36": jugadores_19_36
    }

def obtener_premios_especiales():
    con = db()

    premios = con.execute("""
        SELECT
            p.id,
            p.tipo,
            p.categoria,
            p.jugador_id,
            j.nombre AS jugador_nombre,
            j.handicap AS jugador_handicap
        FROM premios_especiales p
        LEFT JOIN jugadores j
            ON p.jugador_id = j.id
        ORDER BY p.id ASC
    """).fetchall()

    con.close()

    resultado = {}

    for p in premios:
        clave = f"{p['tipo']} {p['categoria']}"
        resultado[clave] = dict(p)

    return resultado

def existe_comodin(equipo):
    con = db()

    existe = con.execute("""
        SELECT id
        FROM jugadores_equipos
        WHERE equipo = ?
        AND comodin = 1
    """, (equipo,)).fetchone()

    con.close()

    return existe is not None

def calcular_resultado_match(puntos_team22, puntos_aguilas, comodin_team22, comodin_aguilas):
    comodin_en_juego = comodin_team22 or comodin_aguilas

    if puntos_team22 > puntos_aguilas:
        if comodin_en_juego:
            return 2, 0, "Gana Team 22"
        return 1, 0, "Gana Team 22"

    if puntos_aguilas > puntos_team22:
        if comodin_en_juego:
            return 0, 2, "Gana Águilas"
        return 0, 1, "Gana Águilas"

    if comodin_en_juego:
        return 1, 1, "Empate"

    return 0.5, 0.5, "Empate"

def obtener_proximo_numero_match():
    con = db()

    ultimo = con.execute("""
        SELECT MAX(numero_match) AS ultimo
        FROM matches_equipos
    """).fetchone()

    con.close()

    if ultimo and ultimo["ultimo"]:
        return int(ultimo["ultimo"]) + 1

    return 1

def obtener_jugadores_equipos():
    con = db()

    team22 = con.execute("""
        SELECT *
        FROM jugadores_equipos
        WHERE equipo = 'Team 22'
        AND id NOT IN (
            SELECT jugador_team22_id
            FROM matches_equipos
        )
        ORDER BY nombre ASC
    """).fetchall()

    aguilas = con.execute("""
        SELECT *
        FROM jugadores_equipos
        WHERE equipo = 'Águilas'
        AND id NOT IN (
            SELECT jugador_aguilas_id
            FROM matches_equipos
        )
        ORDER BY nombre ASC
    """).fetchall()

    team22_todos = con.execute("""
        SELECT *
        FROM jugadores_equipos
        WHERE equipo = 'Team 22'
        ORDER BY nombre ASC
    """).fetchall()

    aguilas_todos = con.execute("""
        SELECT *
        FROM jugadores_equipos
        WHERE equipo = 'Águilas'
        ORDER BY nombre ASC
    """).fetchall()

    con.close()

    return team22, aguilas, team22_todos, aguilas_todos

def obtener_matches_equipos():
    con = db()

    matches = con.execute("""
        SELECT
            m.id,
            m.numero_match,
            m.modo,

            j22.nombre AS jugador_team22,
            j22.comodin AS comodin_team22,

            j22b.nombre AS jugador_team22_2,
            j22b.comodin AS comodin_team22_2,

            ja.nombre AS jugador_aguilas,
            ja.comodin AS comodin_aguilas,

            jab.nombre AS jugador_aguilas_2,
            jab.comodin AS comodin_aguilas_2,

            m.puntos_partido_team22,
            m.puntos_partido_aguilas,
            m.puntos_tabla_team22,
            m.puntos_tabla_aguilas,
            m.resultado,
            m.resultado_cargado

        FROM matches_equipos m

        JOIN jugadores_equipos j22
            ON m.jugador_team22_id = j22.id

        JOIN jugadores_equipos ja
            ON m.jugador_aguilas_id = ja.id

        LEFT JOIN jugadores_equipos j22b
            ON m.jugador_team22_2_id = j22b.id

        LEFT JOIN jugadores_equipos jab
            ON m.jugador_aguilas_2_id = jab.id

        ORDER BY m.numero_match ASC
    """).fetchall()

    con.close()

    total_dia_team22 = sum(float(m["puntos_tabla_team22"]) for m in matches)
    total_dia_aguilas = sum(float(m["puntos_tabla_aguilas"]) for m in matches)

    puntos_previos = obtener_puntos_previos()

    previos_team22 = puntos_previos["Team 22"]
    previos_aguilas = puntos_previos["Águilas"]

    total_team22 = previos_team22 + total_dia_team22
    total_aguilas = previos_aguilas + total_dia_aguilas

    if total_team22 > total_aguilas:
        ganador = "Gana Team 22"
    elif total_aguilas > total_team22:
        ganador = "Gana Águilas"
    else:
        ganador = "Empate"

    pendientes = [
        m for m in matches
        if int(m["resultado_cargado"]) == 0
    ]

    return {
        "matches": matches,
        "pendientes": pendientes,
        "total_dia_team22": total_dia_team22,
        "total_dia_aguilas": total_dia_aguilas,
        "previos_team22": previos_team22,
        "previos_aguilas": previos_aguilas,
        "total_team22": total_team22,
        "total_aguilas": total_aguilas,
        "ganador": ganador
    }

@app.route("/")
def index():
    config = obtener_configuracion()

    return render_template(
        "index.html",
        categorias=obtener_categorias(),
        general=obtener_general(),
        matches_equipos=obtener_matches_equipos(),
        premios_especiales=obtener_premios_especiales(),
        ranking_anual=obtener_ranking_por_fechas(),
        titulo=config["titulo"],
        subtitulo=config["subtitulo"],
        subtitulo2=config["subtitulo2"],
        logo=config["logo"],
        fondo=config["fondo"]
    )

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        password = request.form["password"]

        if password == ADMIN_PASSWORD:
            session["admin"] = True
            session.pop("colaborador", None)
            return redirect("/admin")

        if password == COLAB_PASSWORD:
            session["colaborador"] = True
            session.pop("admin", None)
            return redirect("/colaborador")

        error = "Contraseña incorrecta"

    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

@app.route("/admin", methods=["GET", "POST"])
def admin():

    if request.method == "POST":
        if not admin_o_colaborador():
            return redirect("/login")
    else:
        if not admin_logueado():
            return redirect("/login")

    error = request.args.get("error")
    ok = request.args.get("ok")

    if request.method == "POST":
        jugador_id = int(request.form["jugador"])
        ida = int(request.form["ida"])
        vuelta = int(request.form["vuelta"])

        con = db()

        jugador = con.execute("""
            SELECT *
            FROM jugadores
            WHERE id = ?
        """, (jugador_id,)).fetchone()

        if not jugador:
            con.close()
            return redirect("/admin?error=jugador_no_existe")

        existe = con.execute("""
            SELECT id
            FROM tarjetas
            WHERE jugador_id = ?
        """, (jugador_id,)).fetchone()

        if existe:
            con.close()
            return redirect("/admin?error=jugador_repetido")

        nombre = jugador["nombre"]
        handicap = jugador["handicap"]
        gross = ida + vuelta
        neto = gross - handicap
        categoria = jugador["categoria_anual"]

        if not categoria:
            categoria = categoria_por_hcp(handicap)

        con.execute("""
            INSERT INTO tarjetas
            (jugador_id, nombre, handicap, ida, vuelta, gross, neto, categoria)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            jugador_id,
            nombre,
            handicap,
            ida,
            vuelta,
            gross,
            neto,
            categoria
        ))

        con.commit()
        socketio.emit("actualizar_tabla")
        con.close()

        return redireccion_post_carga("resultado_cargado")

    con = db()

    jugadores = con.execute("""
        SELECT jugadores.*
        FROM jugadores
        LEFT JOIN tarjetas
        ON jugadores.id = tarjetas.jugador_id
        WHERE tarjetas.jugador_id IS NULL
        ORDER BY jugadores.nombre ASC
    """).fetchall()

    jugadores_usados = con.execute("""
        SELECT jugador_team22_id AS jugador_id
        FROM matches_equipos

        UNION

        SELECT jugador_aguilas_id AS jugador_id
        FROM matches_equipos

        UNION

        SELECT jugador_team22_2_id AS jugador_id
        FROM matches_equipos
        WHERE jugador_team22_2_id IS NOT NULL

        UNION

        SELECT jugador_aguilas_2_id AS jugador_id
        FROM matches_equipos
        WHERE jugador_aguilas_2_id IS NOT NULL
    """).fetchall()

    ids_usados = {
        j["jugador_id"]
        for j in jugadores_usados
        if j["jugador_id"] is not None
    }

    con.close()

    config = obtener_configuracion()

    team22, aguilas, team22_todos, aguilas_todos = obtener_jugadores_equipos()

    team22 = [
        j for j in team22
        if j["id"] not in ids_usados
    ]

    aguilas = [
        j for j in aguilas
        if j["id"] not in ids_usados
    ]

    return render_template(
        "admin.html",
        titulo=config["titulo"],
        subtitulo=config["subtitulo"],
        subtitulo2=config["subtitulo2"],
        logo=config["logo"],
        fondo=config["fondo"],
        jugadores=jugadores,
        categorias=obtener_categorias(),
        general=obtener_general(),
        team22=team22,
        aguilas=aguilas,
        team22_todos=team22_todos,
        aguilas_todos=aguilas_todos,
        matches_equipos=obtener_matches_equipos(),
        proximo_match=obtener_proximo_numero_match(),
        comodin_team22_existe=existe_comodin("Team 22"),
        comodin_aguilas_existe=existe_comodin("Águilas"),
        jugadores_premios=obtener_jugadores_premios(),
        premios_especiales=obtener_premios_especiales(),
        ranking_anual=obtener_ranking_por_fechas(),
        ultima_fecha_ranking=obtener_ultima_fecha_ranking(),
        proxima_fecha_ranking=obtener_proxima_fecha_ranking(),
        error=error,
        ok=ok
    )

@app.route("/colaborador")
def colaborador():
    if not colaborador_logueado():
        return redirect("/login")

    con = db()

    jugadores = con.execute("""
        SELECT jugadores.*
        FROM jugadores
        LEFT JOIN tarjetas
        ON jugadores.id = tarjetas.jugador_id
        WHERE tarjetas.jugador_id IS NULL
        ORDER BY jugadores.nombre ASC
    """).fetchall()

    con.close()

    config = obtener_configuracion()

    return render_template(
        "colaborador.html",
        titulo=config["titulo"],
        jugadores=jugadores,
        matches_equipos=obtener_matches_equipos(),
        jugadores_premios=obtener_jugadores_premios(),
        premios_especiales=obtener_premios_especiales()
    )

@app.route("/guardar_configuracion", methods=["POST"])
def guardar_configuracion():
    if not admin_logueado():
        return redirect("/login")

    titulo = request.form["titulo"].strip()
    subtitulo = request.form["subtitulo"].strip()
    subtitulo2 = request.form["subtitulo2"].strip()
    logo = request.form["logo"].strip()
    fondo = request.form["fondo"].strip()

    if not titulo:
        return redirect("/admin?error=titulo_vacio")

    extensiones_validas = (".png", ".jpg", ".jpeg", ".webp")

    if logo and not logo.lower().endswith(extensiones_validas):
        return redirect("/admin?error=logo_invalido")

    if fondo and not fondo.lower().endswith(extensiones_validas):
        return redirect("/admin?error=fondo_invalido")

    con = db()

    con.execute("""
        UPDATE configuracion
        SET titulo = ?, subtitulo = ?, subtitulo2 = ?, logo = ?, fondo = ?
        WHERE id = 1
    """, (titulo, subtitulo, subtitulo2, logo, fondo))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=configuracion_guardada")

@app.route("/cargar_excel", methods=["POST"])
def cargar_excel():
    if not admin_logueado():
        return redirect("/login")

    archivo = request.files.get("archivo")

    if not archivo:
        return redirect("/admin?error=sin_archivo")

    if not archivo.filename.endswith(".xlsx"):
        return redirect("/admin?error=archivo_invalido")

    ruta = os.path.join(UPLOAD_FOLDER, archivo.filename)
    archivo.save(ruta)

    workbook = load_workbook(ruta)
    hoja = workbook.active

    con = db()

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombre = fila[0]
        handicap = fila[1]

        if not nombre or handicap is None:
            continue

        nombre = str(nombre).strip()
        handicap = int(handicap)

        if handicap < 0 or handicap > 36:
            continue

        categoria_anual = categoria_por_hcp(handicap)

        existe = con.execute("""
            SELECT id
            FROM jugadores
            WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(?))
        """, (nombre,)).fetchone()

        if existe:
            continue

        con.execute("""
            INSERT INTO jugadores (nombre, handicap, categoria_anual)
            VALUES (?, ?, ?)
        """, (
            nombre,
            handicap,
            categoria_anual
        ))

        jugador_id = con.execute("""
            SELECT last_insert_rowid()
        """).fetchone()[0]

        matricula = f"MGT-{jugador_id:04d}"

        con.execute("""
            UPDATE jugadores
            SET matricula = ?
            WHERE id = ?
        """, (
            matricula,
            jugador_id
        ))

    con.commit()
    con.close()

    return redirect("/admin?ok=jugadores_cargados")

@app.route("/agregar_jugador", methods=["POST"])
def agregar_jugador():
    if not admin_logueado():
        return redirect("/login")

    nombre = request.form["nombre"].strip()
    handicap = int(request.form["handicap"])

    if not nombre:
        return redirect("/admin?error=nombre_vacio")

    if handicap < 0 or handicap > 36:
        return redirect("/admin?error=handicap_invalido")

    categoria_anual = categoria_por_hcp(handicap)

    con = db()

    try:
        con.execute("""
            INSERT INTO jugadores (nombre, handicap, categoria_anual)
            VALUES (?, ?, ?)
        """, (nombre, handicap, categoria_anual))

        jugador_id = con.execute("""
            SELECT last_insert_rowid()
        """).fetchone()[0]

        matricula = f"MGT-{jugador_id:04d}"

        con.execute("""
            UPDATE jugadores
            SET matricula = ?
            WHERE id = ?
        """, (matricula, jugador_id))

        con.commit()
        con.close()

        return redirect("/admin?ok=jugador_agregado")

    except sqlite3.IntegrityError:
        con.close()
        return redirect("/admin?error=jugador_ya_existe")

@app.route("/editar_jugador/<int:id>", methods=["POST"])
def editar_jugador(id):
    if not admin_logueado():
        return redirect("/login")

    nombre = request.form["nombre"].strip()
    handicap = int(request.form["handicap"])

    if not nombre:
        return redirect("/admin?error=nombre_vacio")

    if handicap < 0 or handicap > 36:
        return redirect("/admin?error=handicap_invalido")

    categoria = categoria_por_hcp(handicap)

    con = db()

    try:
        con.execute("""
            UPDATE jugadores
            SET nombre = ?, handicap = ?
            WHERE id = ?
        """, (nombre, handicap, id))

        tarjeta = con.execute("""
            SELECT *
            FROM tarjetas
            WHERE jugador_id = ?
        """, (id,)).fetchone()

        if tarjeta:
            gross = tarjeta["ida"] + tarjeta["vuelta"]
            neto = gross - handicap

            con.execute("""
                UPDATE tarjetas
                SET nombre = ?, handicap = ?, gross = ?, neto = ?, categoria = ?
                WHERE jugador_id = ?
            """, (nombre, handicap, gross, neto, categoria, id))

        con.commit()
        socketio.emit("actualizar_tabla")
        con.close()

        return redirect("/admin?ok=jugador_modificado")

    except sqlite3.IntegrityError:
        con.close()
        return redirect("/admin?error=jugador_ya_existe")

@app.route("/editar_resultado/<int:id>", methods=["POST"])
def editar_resultado(id):
    if not admin_logueado():
        return redirect("/login")

    ida = int(request.form["ida"])
    vuelta = int(request.form["vuelta"])

    con = db()

    tarjeta = con.execute("""
        SELECT *
        FROM tarjetas
        WHERE id = ?
    """, (id,)).fetchone()

    if not tarjeta:
        con.close()
        return redirect("/admin?error=resultado_no_existe")

    handicap = tarjeta["handicap"]
    gross = ida + vuelta
    neto = gross - handicap

    con.execute("""
        UPDATE tarjetas
        SET ida = ?, vuelta = ?, gross = ?, neto = ?
        WHERE id = ?
    """, (ida, vuelta, gross, neto, id))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=resultado_modificado")

@app.route("/agregar_jugador_equipo", methods=["POST"])
def agregar_jugador_equipo():
    if not admin_logueado():
        return redirect("/login")

    nombre = request.form["nombre"].strip()
    equipo = request.form["equipo"].strip()
    comodin = 1 if request.form.get("comodin") == "1" else 0

    if not nombre:
        return redirect("/admin?error=nombre_vacio")

    if equipo not in ["Team 22", "Águilas"]:
        return redirect("/admin?error=equipo_invalido")

    if comodin == 1 and existe_comodin(equipo):
        return redirect("/admin?error=comodin_existente")

    con = db()

    con.execute("""
        INSERT INTO jugadores_equipos (nombre, equipo, comodin)
        VALUES (?, ?, ?)
    """, (nombre, equipo, comodin))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=jugador_equipo_agregado")

@app.route("/borrar_jugador_equipo/<int:id>")
def borrar_jugador_equipo(id):
    if not admin_logueado():
        return redirect("/login")

    con = db()

    con.execute("""
        DELETE FROM matches_equipos
        WHERE jugador_team22_id = ?
        OR jugador_aguilas_id = ?
    """, (id, id))

    con.execute("""
        DELETE FROM jugadores_equipos
        WHERE id = ?
    """, (id,))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=jugador_equipo_borrado")

@app.route("/crear_cruce_equipo", methods=["POST"])
@app.route("/agregar_match_equipo", methods=["POST"])
def crear_cruce_equipo():
    if not admin_logueado():
        return redirect("/login")

    numero_match = obtener_proximo_numero_match()

    modo = request.form.get("modo", "individual")

    jugador_team22_id = int(request.form["jugador_team22_id"])
    jugador_aguilas_id = int(request.form["jugador_aguilas_id"])

    jugador_team22_2_id = request.form.get("jugador_team22_2_id")
    jugador_aguilas_2_id = request.form.get("jugador_aguilas_2_id")

    con = db()

    ids_jugadores = [
        jugador_team22_id,
        jugador_aguilas_id
    ]

    if modo == "pareja":
        if not jugador_team22_2_id or not jugador_aguilas_2_id:
            con.close()
            return redirect("/admin?error=pareja_incompleta")

        jugador_team22_2_id = int(jugador_team22_2_id)
        jugador_aguilas_2_id = int(jugador_aguilas_2_id)

        ids_jugadores.extend([
            jugador_team22_2_id,
            jugador_aguilas_2_id
        ])
    else:
        jugador_team22_2_id = None
        jugador_aguilas_2_id = None

    if len(ids_jugadores) != len(set(ids_jugadores)):
        con.close()
        return redirect("/admin?error=jugador_repetido_pareja")

    placeholders = ",".join("?" for _ in ids_jugadores)

    ya_usado = con.execute(f"""
        SELECT id
        FROM matches_equipos
        WHERE jugador_team22_id IN ({placeholders})
        OR jugador_aguilas_id IN ({placeholders})
        OR jugador_team22_2_id IN ({placeholders})
        OR jugador_aguilas_2_id IN ({placeholders})
    """, ids_jugadores * 4).fetchone()

    if ya_usado:
        con.close()
        return redirect("/admin?error=jugador_ya_tiene_match")

    con.execute("""
        INSERT INTO matches_equipos
        (
            numero_match,
            modo,
            jugador_team22_id,
            jugador_team22_2_id,
            jugador_aguilas_id,
            jugador_aguilas_2_id,
            resultado
        )
        VALUES (?, ?, ?, ?, ?, ?, 'Pendiente')
    """, (
        numero_match,
        modo,
        jugador_team22_id,
        jugador_team22_2_id,
        jugador_aguilas_id,
        jugador_aguilas_2_id
    ))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=cruce_equipo_creado")

@app.route("/cargar_resultado_match_equipo", methods=["POST"])
def cargar_resultado_match_equipo():
    if not admin_o_colaborador():
        return redirect("/login")

    match_id = int(request.form["match_id"])
    puntos_partido_team22 = float(request.form["puntos_partido_team22"])
    puntos_partido_aguilas = float(request.form["puntos_partido_aguilas"])

    con = db()

    match = con.execute("""
        SELECT
            m.*,
            j22.comodin AS comodin_team22,
            ja.comodin AS comodin_aguilas
        FROM matches_equipos m
        JOIN jugadores_equipos j22
            ON m.jugador_team22_id = j22.id
        JOIN jugadores_equipos ja
            ON m.jugador_aguilas_id = ja.id
        WHERE m.id = ?
    """, (match_id,)).fetchone()

    if not match:
        con.close()
        return redirect("/admin?error=match_no_existe")

    puntos_tabla_team22, puntos_tabla_aguilas, resultado = calcular_resultado_match(
        puntos_partido_team22,
        puntos_partido_aguilas,
        match["comodin_team22"] == 1,
        match["comodin_aguilas"] == 1
    )

    con.execute("""
        UPDATE matches_equipos
        SET
            puntos_partido_team22 = ?,
            puntos_partido_aguilas = ?,
            puntos_tabla_team22 = ?,
            puntos_tabla_aguilas = ?,
            resultado = ?,
            resultado_cargado = 1
        WHERE id = ?
    """, (
        puntos_partido_team22,
        puntos_partido_aguilas,
        puntos_tabla_team22,
        puntos_tabla_aguilas,
        resultado,
        match_id
    ))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redireccion_post_carga("resultado_match_cargado")

@app.route("/guardar_puntos_previos", methods=["POST"])
def guardar_puntos_previos():
    if not admin_logueado():
        return redirect("/login")

    puntos_team22 = float(request.form["puntos_team22"])
    puntos_aguilas = float(request.form["puntos_aguilas"])

    con = db()

    con.execute("""
        UPDATE puntos_equipos
        SET puntos_previos = ?
        WHERE equipo = 'Team 22'
    """, (puntos_team22,))

    con.execute("""
        UPDATE puntos_equipos
        SET puntos_previos = ?
        WHERE equipo = 'Águilas'
    """, (puntos_aguilas,))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=puntos_previos_guardados")

@app.route("/editar_match_equipo/<int:id>", methods=["POST"])
def editar_match_equipo(id):
    if not admin_logueado():
        return redirect("/login")

    puntos_partido_team22 = float(request.form["puntos_partido_team22"])
    puntos_partido_aguilas = float(request.form["puntos_partido_aguilas"])

    con = db()

    match = con.execute("""
        SELECT
            m.*,
            j22.comodin AS comodin_team22,
            ja.comodin AS comodin_aguilas
        FROM matches_equipos m
        JOIN jugadores_equipos j22
            ON m.jugador_team22_id = j22.id
        JOIN jugadores_equipos ja
            ON m.jugador_aguilas_id = ja.id
        WHERE m.id = ?
    """, (id,)).fetchone()

    if not match:
        con.close()
        return redirect("/admin?error=match_no_existe")

    puntos_tabla_team22, puntos_tabla_aguilas, resultado = calcular_resultado_match(
        puntos_partido_team22,
        puntos_partido_aguilas,
        match["comodin_team22"] == 1,
        match["comodin_aguilas"] == 1
    )

    con.execute("""
        UPDATE matches_equipos
        SET
            puntos_partido_team22 = ?,
            puntos_partido_aguilas = ?,
            puntos_tabla_team22 = ?,
            puntos_tabla_aguilas = ?,
            resultado = ?,
            resultado_cargado = 1
        WHERE id = ?
    """, (
        puntos_partido_team22,
        puntos_partido_aguilas,
        puntos_tabla_team22,
        puntos_tabla_aguilas,
        resultado,
        id
    ))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=match_equipo_modificado")

@app.route("/borrar_match_equipo/<int:id>")
def borrar_match_equipo(id):
    if not admin_logueado():
        return redirect("/login")

    con = db()

    con.execute("""
        DELETE FROM matches_equipos
        WHERE id = ?
    """, (id,))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=match_equipo_borrado")

@app.route("/cerrar_match_play")
def cerrar_match_play():
    if not admin_logueado():
        return redirect("/login")

    matches_data = obtener_matches_equipos()

    total_dia_team22 = float(matches_data["total_dia_team22"])
    total_dia_aguilas = float(matches_data["total_dia_aguilas"])

    con = db()

    con.execute("""
        UPDATE puntos_equipos
        SET puntos_previos = puntos_previos + ?
        WHERE equipo = 'Team 22'
    """, (total_dia_team22,))

    con.execute("""
        UPDATE puntos_equipos
        SET puntos_previos = puntos_previos + ?
        WHERE equipo = 'Águilas'
    """, (total_dia_aguilas,))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=match_play_cerrado")

@app.route("/reset_matches_equipos")
def reset_matches_equipos():
    if not admin_logueado():
        return redirect("/login")

    con = db()

    con.execute("DELETE FROM matches_equipos")

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=matches_equipos_borrados")

@app.route("/cargar_excel_equipos", methods=["POST"])
def cargar_excel_equipos():
    if not admin_logueado():
        return redirect("/login")

    archivo = request.files.get("archivo_equipos")

    if not archivo:
        return redirect("/admin?error=sin_archivo")

    if not archivo.filename.endswith(".xlsx"):
        return redirect("/admin?error=archivo_invalido")

    ruta = os.path.join(UPLOAD_FOLDER, archivo.filename)
    archivo.save(ruta)

    workbook = load_workbook(ruta)
    hoja = workbook.active

    con = db()

    comodin_cargado_team22 = existe_comodin("Team 22")
    comodin_cargado_aguilas = existe_comodin("Águilas")

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombre = fila[0]
        equipo = fila[1]
        comodin_valor = fila[2] if len(fila) > 2 else ""

        if not nombre or not equipo:
            continue

        nombre = str(nombre).strip()
        equipo = str(equipo).strip()

        if equipo.lower() in ["team 22", "team22"]:
            equipo = "Team 22"
        elif equipo.lower() in ["aguilas", "águilas"]:
            equipo = "Águilas"
        else:
            continue

        comodin_texto = str(comodin_valor).strip().lower()

        es_comodin = comodin_texto in ["si", "sí", "s", "1", "true", "verdadero", "x"]

        if es_comodin:
            if equipo == "Team 22":
                if comodin_cargado_team22:
                    es_comodin = False
                else:
                    comodin_cargado_team22 = True

            if equipo == "Águilas":
                if comodin_cargado_aguilas:
                    es_comodin = False
                else:
                    comodin_cargado_aguilas = True

        existe = con.execute("""
            SELECT id
            FROM jugadores_equipos
            WHERE nombre = ?
            AND equipo = ?
        """, (nombre, equipo)).fetchone()

        if existe:
            continue

        con.execute("""
            INSERT INTO jugadores_equipos (nombre, equipo, comodin)
            VALUES (?, ?, ?)
        """, (
            nombre,
            equipo,
            1 if es_comodin else 0
        ))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=jugadores_equipos_cargados")

@app.route("/reset_team22")
def reset_team22():
    if not admin_logueado():
        return redirect("/login")

    con = db()

    jugadores = con.execute("""
        SELECT id
        FROM jugadores_equipos
        WHERE equipo = 'Team 22'
    """).fetchall()

    ids = [j["id"] for j in jugadores]

    if ids:
        placeholders = ",".join("?" for _ in ids)

        con.execute(f"""
            DELETE FROM matches_equipos
            WHERE jugador_team22_id IN ({placeholders})
            OR jugador_aguilas_id IN ({placeholders})
        """, ids + ids)

        con.execute(f"""
            DELETE FROM jugadores_equipos
            WHERE id IN ({placeholders})
        """, ids)

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=team22_borrado")

@app.route("/reset_aguilas")
def reset_aguilas():
    if not admin_logueado():
        return redirect("/login")

    con = db()

    jugadores = con.execute("""
        SELECT id
        FROM jugadores_equipos
        WHERE equipo = 'Águilas'
    """).fetchall()

    ids = [j["id"] for j in jugadores]

    if ids:
        placeholders = ",".join("?" for _ in ids)

        con.execute(f"""
            DELETE FROM matches_equipos
            WHERE jugador_team22_id IN ({placeholders})
            OR jugador_aguilas_id IN ({placeholders})
        """, ids + ids)

        con.execute(f"""
            DELETE FROM jugadores_equipos
            WHERE id IN ({placeholders})
        """, ids)

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=aguilas_borrado")

@app.route("/guardar_premios_especiales", methods=["POST"])
def guardar_premios_especiales():
    if not admin_o_colaborador():
        return redirect("/login")

    premios = {
        1: request.form.get("approach_0_18"),
        2: request.form.get("approach_19_36"),
        3: request.form.get("long_0_18"),
        4: request.form.get("long_19_36")
    }

    con = db()

    for premio_id, jugador_id in premios.items():

        # Si queda vacío, vuelve a pendiente
        if not jugador_id:
            con.execute("""
                UPDATE premios_especiales
                SET jugador_id = NULL
                WHERE id = ?
            """, (premio_id,))
            continue

        jugador = con.execute("""
            SELECT *
            FROM jugadores
            WHERE id = ?
        """, (jugador_id,)).fetchone()

        if not jugador:
            continue

        hcp = int(jugador["handicap"])

        if premio_id in [1, 3] and not (0 <= hcp <= 18):
            continue

        if premio_id in [2, 4] and not (19 <= hcp <= 36):
            continue

        con.execute("""
            UPDATE premios_especiales
            SET jugador_id = ?
            WHERE id = ?
        """, (jugador_id, premio_id))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redireccion_post_carga("premios_guardados")

@app.route("/cerrar_fecha_ranking")
def cerrar_fecha_ranking():
    if not admin_logueado():
        return redirect("/login")

    numero_fecha = obtener_proxima_fecha_ranking()

    categorias = obtener_categorias()
    general = obtener_general()

    con = db()

    for categoria, jugadores in categorias.items():
        for jugador in jugadores:
            con.execute("""
                INSERT INTO ranking_fechas
                (jugador_id, tipo_ranking, numero_fecha, puntos)
                VALUES (?, 'categoria', ?, ?)
                ON CONFLICT(jugador_id, tipo_ranking, numero_fecha)
                DO UPDATE SET puntos = excluded.puntos
            """, (
                jugador["jugador_id"],
                numero_fecha,
                jugador["puntos"]
            ))

    for jugador in general:
        con.execute("""
            INSERT INTO ranking_fechas
            (jugador_id, tipo_ranking, numero_fecha, puntos)
            VALUES (?, 'general', ?, ?)
            ON CONFLICT(jugador_id, tipo_ranking, numero_fecha)
            DO UPDATE SET puntos = excluded.puntos
        """, (
            jugador["jugador_id"],
            numero_fecha,
            jugador["puntos"]
        ))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect(f"/admin?ok=fecha_{numero_fecha}_cerrada")

@app.route("/importar_ranking_fecha", methods=["POST"])
def importar_ranking_fecha():
    if not admin_logueado():
        return redirect("/login")

    archivo = request.files.get("archivo_ranking")
    tipo_ranking = request.form.get("tipo_ranking")
    numero_fecha = int(request.form.get("numero_fecha"))

    if not archivo:
        return redirect("/admin?error=sin_archivo")

    if not archivo.filename.endswith(".xlsx"):
        return redirect("/admin?error=archivo_invalido")

    if tipo_ranking not in ["categoria", "general"]:
        return redirect("/admin?error=tipo_ranking_invalido")

    ruta = os.path.join(UPLOAD_FOLDER, archivo.filename)
    archivo.save(ruta)

    workbook = load_workbook(ruta, data_only=True)
    con = db()

    hojas = ["Ranking"]

    for nombre_hoja in hojas:
        if nombre_hoja not in workbook.sheetnames:
            continue

        hoja = workbook[nombre_hoja]

        encabezados = {}

        for col in range(1, hoja.max_column + 1):
            valor = hoja.cell(row=1, column=col).value

            if valor:
                encabezados[str(valor).strip().lower()] = col

        col_matricula = encabezados.get("matricula")
        col_puntos = encabezados.get("puntos")

        if not col_matricula or not col_puntos:
            continue

        for fila in range(2, hoja.max_row + 1):
            matricula = hoja.cell(row=fila, column=col_matricula).value
            puntos = hoja.cell(row=fila, column=col_puntos).value

            if not matricula or puntos is None:
                continue

            matricula = str(matricula).strip()

            jugador = con.execute("""
                SELECT id
                FROM jugadores
                WHERE matricula = ?
            """, (matricula,)).fetchone()

            if not jugador:
                continue

            con.execute("""
                INSERT INTO ranking_fechas
                (jugador_id, tipo_ranking, numero_fecha, puntos)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(jugador_id, tipo_ranking, numero_fecha)
                DO UPDATE SET puntos = excluded.puntos
            """, (
                jugador["id"],
                tipo_ranking,
                numero_fecha,
                float(puntos)
            ))

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=ranking_importado")

@app.route("/importar_categorias_anuales", methods=["POST"])
def importar_categorias_anuales():

    if not admin_logueado():
        return redirect("/login")

    archivo = request.files.get("archivo_categorias")

    if not archivo:
        return redirect("/admin?error=sin_archivo")

    ruta = os.path.join(UPLOAD_FOLDER, archivo.filename)
    archivo.save(ruta)

    wb = load_workbook(ruta, data_only=True)

    con = db()

    categorias = {
        "0 a 12": "0 a 12",
        "13 a 22": "13 a 22",
        "23 a 36": "23 a 36"
    }

    for nombre_hoja, categoria in categorias.items():

        if nombre_hoja not in wb.sheetnames:
            continue

        hoja = wb[nombre_hoja]

        for fila in range(2, hoja.max_row + 1):

            matricula = hoja.cell(row=fila, column=1).value

            if not matricula:
                continue

            matricula = str(matricula).strip()

            con.execute("""
                UPDATE jugadores
                SET categoria_anual = ?
                WHERE matricula = ?
            """, (
                categoria,
                matricula
            ))

    con.commit()
    con.close()

    return redirect("/admin?ok=categorias_importadas")

@app.route("/descargar_plantilla_ranking")
def descargar_plantilla_ranking():
    if not admin_logueado():
        return redirect("/login")

    con = db()

    jugadores = con.execute("""
        SELECT matricula, nombre, handicap
        FROM jugadores
        ORDER BY nombre ASC
    """).fetchall()

    con.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    encabezados = [
        "Matricula",
        "Nombre",
        "HCP",
        "Puntos"
    ]

    ws.append(encabezados)

    for celda in ws[1]:
        celda.font = Font(bold=True)

    for j in jugadores:
        ws.append([
            j["matricula"],
            j["nombre"],
            j["handicap"],
            ""
        ])

    for columna in ws.columns:
        max_length = 0
        letra = columna[0].column_letter

        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))

        ws.column_dimensions[letra].width = max_length + 3

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="plantilla_ranking_anual.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/cargar_jugadores_categoria", methods=["POST"])
def cargar_jugadores_categoria():
    if not admin_logueado():
        return redirect("/login")

    categoria_anual = request.form["categoria_anual"]
    archivo = request.files.get("archivo_categoria")

    if not archivo:
        return redirect("/admin?error=sin_archivo")

    if not archivo.filename.endswith(".xlsx"):
        return redirect("/admin?error=archivo_invalido")

    ruta = os.path.join(UPLOAD_FOLDER, archivo.filename)
    archivo.save(ruta)

    workbook = load_workbook(ruta, data_only=True)
    hoja = workbook.active

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombre = fila[0]
        handicap = fila[1]

        if not nombre or handicap is None:
            continue

        nombre = str(nombre).strip()
        handicap = int(handicap)

        jugador_id = crear_o_actualizar_jugador_ranking(
            nombre,
            handicap,
            categoria_anual
        )

        # Columnas esperadas:
        # Nombre | HCP | F1 Cat | F2 Cat | F3 Cat | F1 Gen | F2 Gen | F3 Gen
        guardar_punto_ranking(jugador_id, "categoria", 1, fila[2] if len(fila) > 2 else None)
        guardar_punto_ranking(jugador_id, "categoria", 2, fila[3] if len(fila) > 3 else None)
        guardar_punto_ranking(jugador_id, "categoria", 3, fila[4] if len(fila) > 4 else None)

        guardar_punto_ranking(jugador_id, "general", 1, fila[5] if len(fila) > 5 else None)
        guardar_punto_ranking(jugador_id, "general", 2, fila[6] if len(fila) > 6 else None)
        guardar_punto_ranking(jugador_id, "general", 3, fila[7] if len(fila) > 7 else None)

    socketio.emit("actualizar_tabla")

    return redirect("/admin?ok=jugadores_categoria_cargados")

@app.route("/agregar_jugador_ranking_manual", methods=["POST"])
def agregar_jugador_ranking_manual():
    if not admin_logueado():
        return redirect("/login")

    nombre = request.form["nombre"].strip()
    handicap = int(request.form["handicap"])
    categoria_anual = request.form["categoria_anual"]

    if not nombre:
        return redirect("/admin?error=nombre_vacio")

    jugador_id = crear_o_actualizar_jugador_ranking(
        nombre,
        handicap,
        categoria_anual
    )

    guardar_punto_ranking(jugador_id, "categoria", 1, request.form.get("f1_cat"))
    guardar_punto_ranking(jugador_id, "categoria", 2, request.form.get("f2_cat"))
    guardar_punto_ranking(jugador_id, "categoria", 3, request.form.get("f3_cat"))

    guardar_punto_ranking(jugador_id, "general", 1, request.form.get("f1_gen"))
    guardar_punto_ranking(jugador_id, "general", 2, request.form.get("f2_gen"))
    guardar_punto_ranking(jugador_id, "general", 3, request.form.get("f3_gen"))

    socketio.emit("actualizar_tabla")

    return redirect("/admin?ok=jugador_ranking_agregado")

@app.route("/descargar_backup_db")
def descargar_backup_db():
    if not admin_logueado():
        return redirect("/login")

    fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")

    return send_file(
        DB_PATH,
        as_attachment=True,
        download_name=f"backup_scores_{fecha}.db"
    )

@app.route("/restaurar_backup", methods=["POST"])
def restaurar_backup():

    if not admin_logueado():
        return redirect("/login")

    archivo = request.files.get("backup")

    if not archivo:
        return redirect("/admin?error=sin_backup")

    archivo.save("/var/data/scores_nuevo.db")

    return redirect("/admin?ok=backup_subido")

@app.route("/descargar_resultados")
def descargar_resultados():
    if not admin_logueado():
        return redirect("/login")

    categorias = obtener_categorias()

    wb = Workbook()
    wb.remove(wb.active)

    for categoria, jugadores in categorias.items():
        ws = wb.create_sheet(title=f"Cat {categoria}")

        encabezados = [
            "Puesto",
            "Nombre",
            "Handicap",
            "Ida",
            "Vuelta",
            "Gross",
            "Neto",
            "Puntos",
            "Categoría"
        ]

        ws.append(encabezados)

        for celda in ws[1]:
            celda.font = Font(bold=True)

        for i, j in enumerate(jugadores, start=1):
            ws.append([
                i,
                j["nombre"],
                j["handicap"],
                j["ida"],
                j["vuelta"],
                j["gross"],
                j["neto"],
                j["puntos"],
                j["categoria"]
            ])

        for columna in ws.columns:
            max_length = 0
            letra = columna[0].column_letter

            for celda in columna:
                if celda.value:
                    max_length = max(max_length, len(str(celda.value)))

            ws.column_dimensions[letra].width = max_length + 3

    ws = wb.create_sheet(title="General")

    encabezados = [
        "Puesto",
        "Nombre",
        "Categoría",
        "Neto",
        "Puntos"
    ]

    ws.append(encabezados)

    for celda in ws[1]:
        celda.font = Font(bold=True)

    for i, j in enumerate(obtener_general(), start=1):
        ws.append([
            i,
            j["nombre"],
            j["categoria"],
            j["neto"],
            j["puntos"]
        ])

    for columna in ws.columns:
        max_length = 0
        letra = columna[0].column_letter

        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))

        ws.column_dimensions[letra].width = max_length + 3
        
    ws = wb.create_sheet(title="Match Play")

    encabezados = [
        "Match",
        "Team 22",
        "Pts Partido Team 22",
        "Águilas",
        "Pts Partido Águilas",
        "Puntos Team 22",
        "Puntos Águilas",
        "Resultado"
    ]

    ws.append(encabezados)

    for celda in ws[1]:
        celda.font = Font(bold=True)

    matches_data = obtener_matches_equipos()

    for m in matches_data["matches"]:
        jugador_team22 = m["jugador_team22"]
        if m["comodin_team22"] == 1:
            jugador_team22 += " ⭐"

        jugador_aguilas = m["jugador_aguilas"]
        if m["comodin_aguilas"] == 1:
            jugador_aguilas += " ⭐"

        ws.append([
            m["numero_match"],
            jugador_team22,
            m["puntos_partido_team22"],
            jugador_aguilas,
            m["puntos_partido_aguilas"],
            m["puntos_tabla_team22"],
            m["puntos_tabla_aguilas"],
            m["resultado"]
        ])

    ws.append([])

    ws.append([
        "",
        "TOTAL TEAM 22",
        "",
        "TOTAL ÁGUILAS",
        "",
        matches_data["total_team22"],
        matches_data["total_aguilas"],
        matches_data["ganador"]
    ])

    for columna in ws.columns:
        max_length = 0
        letra = columna[0].column_letter

        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))

        ws.column_dimensions[letra].width = max_length + 3
    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="resultados_torneo_golf.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@app.route("/borrar/<int:id>")
def borrar(id):
    if not admin_logueado():
        return redirect("/login")

    con = db()
    con.execute("DELETE FROM tarjetas WHERE id = ?", (id,))
    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=resultado_borrado")

@app.route("/borrar_jugador/<int:id>")
def borrar_jugador(id):
    if not admin_logueado():
        return redirect("/login")

    con = db()
    con.execute("DELETE FROM tarjetas WHERE jugador_id = ?", (id,))
    con.execute("DELETE FROM jugadores WHERE id = ?", (id,))
    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=jugador_borrado")

@app.route("/limpiar_fecha_actual")
def limpiar_fecha_actual():
    if not admin_logueado():
        return redirect("/login")

    con = db()

    con.execute("DELETE FROM tarjetas")
    con.execute("DELETE FROM matches_equipos")

    con.execute("""
        UPDATE premios_especiales
        SET jugador_id = NULL
    """)

    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=fecha_limpiada")

@app.route("/test_ranking")
def test_ranking():
    con = db()

    cantidad = con.execute("""
        SELECT COUNT(*)
        FROM ranking_anual
    """).fetchone()[0]

    con.close()

    return f"Jugadores en ranking anual: {cantidad}"

@app.route("/reset_jugadores")
def reset_jugadores():
    if not admin_logueado():
        return redirect("/login")

    con = db()
    con.execute("DELETE FROM tarjetas")
    con.execute("DELETE FROM jugadores")
    con.commit()
    socketio.emit("actualizar_tabla")
    con.close()

    return redirect("/admin?ok=todo_borrado")

@app.route("/confirmar_reset_total", methods=["GET", "POST"])
def confirmar_reset_total():

    if not admin_logueado():
        return redirect("/login")

    if request.method == "POST":

        confirmacion = request.form.get("confirmacion", "").strip()

        if confirmacion != "BORRAR TODO":
            return render_template(
                "confirmar_reset_total.html",
                error="Debés escribir exactamente: BORRAR TODO"
            )

        return redirect("/reset_todo")

    return render_template("confirmar_reset_total.html")

@app.route("/reset_todo")
def reset_todo():
    if not admin_logueado():
        return redirect("/login")

    con = db()

    con.execute("DELETE FROM tarjetas")
    con.execute("DELETE FROM jugadores")
    con.execute("DELETE FROM jugadores_equipos")
    con.execute("DELETE FROM matches_equipos")
    con.execute("DELETE FROM ranking_fechas")
    con.execute("DELETE FROM puntos_equipos")
    con.execute("DELETE FROM premios_especiales")
    con.execute("DELETE FROM configuracion")

    con.commit()
    con.close()

    init_db()

    socketio.emit("actualizar_tabla")

    return redirect("/admin?ok=reset_total")

if __name__ == "__main__":
    init_db()
    asignar_matriculas()
    inicializar_ranking_anual()
    socketio.run(app, host="127.0.0.1", port=5000, debug=True)
else:
    init_db()
    asignar_matriculas()
    inicializar_ranking_anual()
